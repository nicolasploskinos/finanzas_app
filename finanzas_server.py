import os
import csv
import io
import re
import time
import random
import unicodedata
import calendar
import hmac
import hashlib
import requests as req
from google import genai
from google.genai import types as genai_types
from urllib.parse import quote
from flask import Flask, jsonify, request, render_template, send_from_directory, session, redirect, Response
from flask_cors import CORS
from supabase import create_client
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import date, timedelta, datetime, timezone
from dotenv import load_dotenv

load_dotenv()

_inflacion_cache = {"data": None, "ts": 0}
_cotiz_cache = {"data": None, "ts": 0}
_wa_codigos = {}       # codigo de vinculación -> (user_id, expira_ts)
_wa_procesados = set() # wamids ya procesados, para ignorar reintentos de Meta

def _hoy_ar():
    # El servidor corre en UTC; Buenos Aires es UTC-3 todo el año (sin horario de verano).
    return (datetime.now(timezone.utc) - timedelta(hours=3)).date()

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ["SECRET_KEY"]
app.permanent_session_lifetime = timedelta(days=30)
app.config["SESSION_COOKIE_SECURE"]   = True
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024  # tope global de subida (el import de CSV ya recorta a 2MB aparte)

db = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])

_login_intentos = {}           # username -> [timestamps de intentos fallidos]
_LOGIN_MAX_INTENTOS = 5
_LOGIN_VENTANA_SEGUNDOS = 15 * 60

def _login_bloqueado(username):
    ahora = time.time()
    intentos = [ts for ts in _login_intentos.get(username, []) if ahora - ts < _LOGIN_VENTANA_SEGUNDOS]
    _login_intentos[username] = intentos
    return len(intentos) >= _LOGIN_MAX_INTENTOS

def _registrar_intento_fallido(username):
    _login_intentos.setdefault(username, []).append(time.time())

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect("/finanzas/login")
        return f(*args, **kwargs)
    return decorated

# ── Páginas ───────────────────────────────────────────────────────────────────

@app.after_request
def _no_cache_html(resp):
    if resp.mimetype == "text/html":
        resp.headers["Cache-Control"] = "no-store"
    return resp

@app.route("/")
def landing():
    return render_template("landing.html")

@app.route("/finanzas")
@login_required
def index():
    user = db.table("usuarios").select("plan,trial_expira").eq("id", session["user_id"]).execute().data
    is_pro = False
    is_trial = False
    trial_dias = 0
    if user:
        plan = user[0].get("plan")
        trial_expira = user[0].get("trial_expira")
        if plan == "pro":
            is_pro = True
        elif plan == "trial" and trial_expira:
            expira = datetime.fromisoformat(trial_expira.replace("Z", "+00:00"))
            if expira > datetime.now(timezone.utc):
                is_pro = True
                is_trial = True
                trial_dias = max(1, (expira - datetime.now(timezone.utc)).days + 1)
    return render_template("finanzas.html", username=session["username"], is_pro=is_pro, is_trial=is_trial, trial_dias=trial_dias)

@app.route("/finanzas/viajes")
@login_required
def viajes_page():
    return render_template("viajes.html", username=session["username"], is_pro=_es_pro(session["user_id"]))

@app.route("/finanzas/login")
def login_page():
    if "user_id" in session:
        return redirect("/finanzas")
    return render_template("auth.html")

@app.route("/finanzas/logout")
def logout():
    session.clear()
    return redirect("/finanzas/login")

@app.route("/finanzas/manifest.json")
def manifest():
    return send_from_directory("static", "finanzas_manifest.json", mimetype="application/manifest+json")

@app.route("/finanzas/sw.js")
def sw():
    return send_from_directory("static", "finanzas_sw.js", mimetype="application/javascript")

# ── Auth API ──────────────────────────────────────────────────────────────────

@app.route("/api/finanzas/login", methods=["POST"])
def login():
    data     = request.get_json()
    username = (data.get("username") or "").strip()
    password =  data.get("password") or ""

    if _login_bloqueado(username):
        return jsonify({"ok": False, "error": "too_many_attempts"}), 429

    res = db.table("usuarios").select("*").eq("username", username).execute()
    if not res.data or not check_password_hash(res.data[0]["password_hash"], password):
        _registrar_intento_fallido(username)
        return jsonify({"ok": False, "error": "invalid_credentials"}), 401

    user = res.data[0]
    session.permanent = True
    session["user_id"]  = user["id"]
    session["username"] = user["username"]
    _login_intentos.pop(username, None)
    return jsonify({"ok": True})

@app.route("/api/finanzas/register", methods=["POST"])
def register():
    data     = request.get_json()
    username = (data.get("username") or "").strip()
    password =  data.get("password") or ""

    if not username or not password:
        return jsonify({"ok": False, "error": "missing_fields"}), 400

    if db.table("usuarios").select("id").eq("username", username).execute().data:
        return jsonify({"ok": False, "error": "username_taken"}), 400

    res = db.table("usuarios").insert({
        "username":      username,
        "password_hash": generate_password_hash(password),
        "verificado":    True,
        "plan":          "trial",
        "trial_expira":  (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
    }).execute()

    user = res.data[0]
    db.table("transacciones").update({"user_id": user["id"]}).is_("user_id", "null").execute()

    session.permanent = True
    session["user_id"]  = user["id"]
    session["username"] = user["username"]
    return jsonify({"ok": True}), 201

# ── Cotizaciones API ──────────────────────────────────────────────────────────

def _obtener_cotizaciones():
    if _cotiz_cache["data"] and time.time() - _cotiz_cache["ts"] < 600:
        return _cotiz_cache["data"]
    try:
        r = req.get("https://api.bluelytics.com.ar/v2/latest", timeout=5)
        data = r.json()
        cotiz = {
            "USD": round(data["oficial"]["value_sell"], 2),
            "EUR": round(data["oficial_euro"]["value_sell"], 2),
        }
    except Exception:
        cotiz = {"USD": None, "EUR": None}
    _cotiz_cache["data"] = cotiz
    _cotiz_cache["ts"] = time.time()
    return cotiz

def _convertir_moneda(monto, origen, destino, cotiz):
    if origen == destino:
        return monto
    ars = monto if origen == "ARS" else (monto * cotiz[origen] if cotiz.get(origen) else None)
    if ars is None:
        return None
    if destino == "ARS":
        return ars
    return ars / cotiz[destino] if cotiz.get(destino) else None

@app.route("/api/finanzas/cotizaciones")
def cotizaciones():
    return jsonify(_obtener_cotizaciones())

# ── Plan helpers ─────────────────────────────────────────────────────────────

def _es_pro(user_id):
    res = db.table("usuarios").select("plan,trial_expira").eq("id", user_id).execute()
    if not res.data:
        return False
    u = res.data[0]
    if u.get("plan") == "pro":
        return True
    if u.get("plan") == "trial" and u.get("trial_expira"):
        expira = datetime.fromisoformat(u["trial_expira"].replace("Z", "+00:00"))
        return expira > datetime.now(timezone.utc)
    return False

# ── Recurrentes helpers ───────────────────────────────────────────────────────

def _siguiente_fecha(fecha_str, frecuencia):
    d = date.fromisoformat(fecha_str)
    if frecuencia == "semanal":
        return (d + timedelta(weeks=1)).isoformat()
    mes  = d.month + 1 if d.month < 12 else 1
    anio = d.year      if d.month < 12 else d.year + 1
    return date(anio, mes, min(d.day, calendar.monthrange(anio, mes)[1])).isoformat()

def _procesar_recurrentes(user_id):
    hoy = _hoy_ar().isoformat()
    pendientes = db.table("recurrentes").select("*").eq("user_id", user_id).eq("activo", True).lte("proxima_fecha", hoy).execute().data
    for r in pendientes:
        prox = r["proxima_fecha"]
        while prox <= hoy:
            db.table("transacciones").insert({
                "tipo": r["tipo"], "monto": r["monto"], "fecha": prox,
                "categoria": r["categoria"], "descripcion": r["descripcion"],
                "moneda": r["moneda"], "user_id": user_id,
            }).execute()
            prox = _siguiente_fecha(prox, r["frecuencia"])
        db.table("recurrentes").update({"proxima_fecha": prox}).eq("id", r["id"]).execute()

# ── Transacciones API ─────────────────────────────────────────────────────────

@app.route("/api/finanzas", methods=["GET"])
@login_required
def listar():
    try:
        _procesar_recurrentes(session["user_id"])
    except Exception:
        pass
    es_pro = _es_pro(session["user_id"])
    query = db.table("transacciones").select("*").eq("user_id", session["user_id"])
    if not es_pro:
        desde = (_hoy_ar() - timedelta(days=90)).isoformat()
        query = query.gte("fecha", desde)
    res = query.order("fecha", desc=True).execute()
    return jsonify(res.data)

def _insertar_transaccion(user_id, tipo, monto, fecha, categoria="", descripcion="", moneda="ARS", viaje_id=None):
    if not _es_pro(user_id):
        inicio_mes = _hoy_ar().replace(day=1).isoformat()
        count = len(db.table("transacciones").select("id").eq("user_id", user_id).gte("fecha", inicio_mes).execute().data)
        if count >= 50:
            return False, "limite_pro"
    payload = {
        "tipo": tipo, "monto": float(monto), "fecha": fecha,
        "categoria": categoria or "", "descripcion": descripcion or "",
        "moneda": moneda or "ARS", "user_id": user_id, "viaje_id": viaje_id,
    }
    res = db.table("transacciones").insert(payload).execute()
    return True, res.data[0]

@app.route("/api/finanzas", methods=["POST"])
@login_required
def agregar():
    t = request.get_json()
    ok, resultado = _insertar_transaccion(
        session["user_id"], t["tipo"], t["monto"], t["fecha"],
        t.get("categoria", ""), t.get("descripcion", ""), t.get("moneda", "ARS"),
        t.get("viaje_id"),
    )
    if not ok:
        return jsonify({"ok": False, "error": resultado}), 403
    return jsonify(resultado), 201

@app.route("/api/finanzas/export", methods=["GET"])
@login_required
def exportar():
    if not _es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    res = db.table("transacciones").select("*").eq("user_id", session["user_id"]).order("fecha", desc=True).execute()
    output = io.StringIO()
    output.write("sep=;\n")
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Fecha", "Tipo", "Monto", "Moneda", "Categoria", "Descripcion"])
    for t in res.data:
        writer.writerow([
            t["fecha"], t["tipo"], t["monto"],
            t.get("moneda","ARS"), t.get("categoria",""), t.get("descripcion",""),
        ])
    encoded = output.getvalue().encode("utf-8-sig")
    return Response(encoded, mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=finanzas.csv"})

@app.route("/api/finanzas/export/excel", methods=["GET"])
@login_required
def exportar_excel():
    if not _es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    res = db.table("transacciones").select("*").eq("user_id", session["user_id"]).order("fecha", desc=True).execute()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    headers = ["Fecha", "Tipo", "Monto", "Moneda", "Categoría", "Descripción"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for t in res.data:
        ws.append([
            t["fecha"], t["tipo"], float(t["monto"]),
            t.get("moneda", "ARS"), t.get("categoria", ""), t.get("descripcion", ""),
        ])
    for col, width in zip("ABCDEF", (12, 10, 14, 8, 20, 34)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finanzas.xlsx"},
    )

@app.route("/api/finanzas/export/pdf", methods=["GET"])
@login_required
def exportar_pdf():
    if not _es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    res = db.table("transacciones").select("*").eq("user_id", session["user_id"]).order("fecha", desc=True).execute()

    from fpdf import FPDF

    def _safe(s):
        return (s or "").encode("latin-1", "replace").decode("latin-1")

    total_gastos = sum(float(t["monto"]) for t in res.data if t["tipo"] == "Gasto")
    total_ingresos = sum(float(t["monto"]) for t in res.data if t["tipo"] == "Ingreso")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _safe("Finanzas - Historial de transacciones"), ln=1)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, _safe(f"Generado el {_hoy_ar().isoformat()} - Usuario: {session['username']}"), ln=1)
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, _safe(f"Total ingresos: {total_ingresos:,.2f}   Total gastos: {total_gastos:,.2f}"), ln=1)
    pdf.ln(4)

    col_widths = [22, 18, 26, 14, 40, 68]
    headers = ["Fecha", "Tipo", "Monto", "Moneda", "Categoria", "Descripcion"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(22, 163, 74)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 7, h, border=1, fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for t in res.data:
        fila = [
            t["fecha"], t["tipo"], f'{float(t["monto"]):,.2f}',
            t.get("moneda", "ARS"), (t.get("categoria") or "")[:22], (t.get("descripcion") or "")[:45],
        ]
        for w, val in zip(col_widths, fila):
            pdf.cell(w, 6, _safe(str(val)), border=1)
        pdf.ln()
        if pdf.get_y() > 270:
            pdf.add_page()

    salida = bytes(pdf.output())
    return Response(
        salida, mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=finanzas.pdf"},
    )

# ── Importar movimientos (banco / billetera virtual) ──────────────────────────

def _norm_header(s):
    s = (s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def _buscar_col(headers, candidatos):
    for i, h in enumerate(headers):
        if any(c in h for c in candidatos):
            return i
    return None

def _parse_fecha_import(s):
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None

def _parse_monto_import(s):
    s = (s or "").strip()
    if not s:
        return None
    neg = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    s = re.sub(r"[^\d,.\-]", "", s).lstrip("-")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        partes = s.split(",")
        if len(partes[-1]) == 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    if not s:
        return None
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val

@app.route("/api/finanzas/import/preview", methods=["POST"])
@login_required
def import_preview():
    if not _es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403

    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename.lower().endswith(".csv"):
        return jsonify({"ok": False, "error": "invalid_file_type"}), 400

    raw = archivo.read(2_000_000)
    texto = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        return jsonify({"ok": False, "error": "cant_read_file"}), 400

    muestra = texto[:2000]
    try:
        delim = csv.Sniffer().sniff(muestra, delimiters=",;\t").delimiter
    except csv.Error:
        delim = ";" if muestra.count(";") > muestra.count(",") else ","

    filas = list(csv.reader(io.StringIO(texto), delimiter=delim))
    if len(filas) < 2:
        return jsonify({"ok": False, "error": "empty_file"}), 400

    headers = [_norm_header(h) for h in filas[0]]
    idx_fecha = _buscar_col(headers, ["fecha", "date"])
    idx_monto = _buscar_col(headers, ["importe", "monto", "valor", "amount", "total"])
    idx_desc  = _buscar_col(headers, ["descripcion", "concepto", "detalle", "description", "glosa"])
    idx_debe  = _buscar_col(headers, ["debe", "egreso", "cargo", "debito"])
    idx_haber = _buscar_col(headers, ["haber", "ingreso", "abono", "credito"])

    if idx_fecha is None or (idx_monto is None and (idx_debe is None or idx_haber is None)):
        return jsonify({
            "ok": False,
            "error": "unrecognized_columns",
        }), 400

    transacciones = []
    errores = 0
    for fila in filas[1:]:
        if not fila or all(not c.strip() for c in fila):
            continue
        try:
            fecha = _parse_fecha_import(fila[idx_fecha]) if idx_fecha < len(fila) else None
            desc  = fila[idx_desc].strip() if idx_desc is not None and idx_desc < len(fila) else ""

            if idx_monto is not None:
                monto_raw = _parse_monto_import(fila[idx_monto]) if idx_monto < len(fila) else None
                if not monto_raw:
                    errores += 1
                    continue
                tipo, monto = ("Gasto" if monto_raw < 0 else "Ingreso"), abs(monto_raw)
            else:
                debe  = abs(_parse_monto_import(fila[idx_debe])  or 0) if idx_debe  < len(fila) else 0
                haber = abs(_parse_monto_import(fila[idx_haber]) or 0) if idx_haber < len(fila) else 0
                if haber > 0:
                    tipo, monto = "Ingreso", haber
                elif debe > 0:
                    tipo, monto = "Gasto", debe
                else:
                    errores += 1
                    continue

            if not fecha:
                errores += 1
                continue

            transacciones.append({
                "fecha": fecha, "tipo": tipo, "monto": round(monto, 2),
                "descripcion": desc, "categoria": "", "moneda": "ARS",
            })
        except Exception:
            errores += 1

        if len(transacciones) >= 1000:
            break

    return jsonify({"ok": True, "transacciones": transacciones, "errores": errores})

@app.route("/api/finanzas/import/confirm", methods=["POST"])
@login_required
def import_confirm():
    if not _es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403

    data = request.get_json() or {}
    filas = data.get("transacciones") or []
    if not isinstance(filas, list) or not filas:
        return jsonify({"ok": False, "error": "nothing_to_import"}), 400
    if len(filas) > 1000:
        return jsonify({"ok": False, "error": "too_many_transactions"}), 400

    payload = []
    for t in filas:
        try:
            payload.append({
                "tipo":        "Ingreso" if t.get("tipo") == "Ingreso" else "Gasto",
                "monto":       abs(float(t["monto"])),
                "fecha":       date.fromisoformat(t["fecha"]).isoformat(),
                "categoria":   (t.get("categoria") or "").strip()[:60],
                "descripcion": (t.get("descripcion") or "").strip()[:200],
                "moneda":      t.get("moneda") if t.get("moneda") in ("ARS", "USD", "EUR") else "ARS",
                "user_id":     session["user_id"],
            })
        except (KeyError, ValueError, TypeError):
            continue

    if not payload:
        return jsonify({"ok": False, "error": "nothing_valid_to_import"}), 400

    insertados = 0
    for i in range(0, len(payload), 500):
        lote = payload[i:i + 500]
        db.table("transacciones").insert(lote).execute()
        insertados += len(lote)

    return jsonify({"ok": True, "insertados": insertados})

# ── Bot de WhatsApp ────────────────────────────────────────────────────────────

_WA_PALABRAS_INGRESO = {
    "ingreso", "ingrese", "ingresaron",
    "cobre", "cobro",
    "deposito", "depositaron",
    "recibi",
    "gane",
    "sueldo",
}
_WA_PALABRAS_DESCARTE = {
    "gaste", "gasto", "pague",
    "cobre", "cobro", "ingreso", "ingrese", "ingresaron", "deposito", "depositaron", "recibi", "gane",
    "en", "de", "del", "la", "el", "los", "las", "me",
    "usd", "dolar", "dolares", "eur", "euro", "euros",
}

def _parse_mensaje_whatsapp(texto):
    desc_match = re.search(r"\(([^)]+)\)", texto)
    descripcion = desc_match.group(1).strip() if desc_match else ""
    if desc_match:
        texto = texto[:desc_match.start()] + texto[desc_match.end():]

    limpio = _norm_header(texto)
    if not limpio:
        return None

    match = re.search(r"\d[\d.,]*", limpio)
    if not match:
        return None
    monto = _parse_monto_import(match.group())
    if not monto:
        return None

    palabras = re.findall(r"[a-z0-9$]+", limpio)
    tipo = "Ingreso" if any(p in _WA_PALABRAS_INGRESO for p in palabras) else "Gasto"

    moneda = "ARS"
    if "usd" in palabras or "dolar" in palabras or "dolares" in palabras or "u$s" in palabras:
        moneda = "USD"
    elif "eur" in palabras or "euro" in palabras or "euros" in palabras:
        moneda = "EUR"

    resto = limpio[:match.start()] + limpio[match.end():]
    cat_palabras = [w for w in re.findall(r"[a-z]+", resto) if w not in _WA_PALABRAS_DESCARTE]
    categoria = " ".join(cat_palabras).strip().capitalize()

    return {"tipo": tipo, "monto": round(abs(monto), 2), "moneda": moneda, "categoria": categoria, "descripcion": descripcion}

def _wa_responder_pregunta(user_id, pregunta):
    if not os.environ.get("GEMINI_API_KEY"):
        return None
    hoy_ar = _hoy_ar()
    desde = (hoy_ar - timedelta(days=180)).isoformat()
    res = (db.table("transacciones").select("tipo,monto,moneda,categoria,descripcion,fecha")
           .eq("user_id", user_id).gte("fecha", desde).order("fecha", desc=True).limit(300).execute())
    if not res.data:
        return "Todavía no tenés movimientos cargados para que te pueda contar algo 🤔"

    lineas = "\n".join(
        f"{t['fecha']} | {t['tipo']} | {t['monto']} {t.get('moneda') or 'ARS'} | "
        f"{t.get('categoria') or 'sin categoría'} | {t.get('descripcion') or ''}"
        for t in res.data
    )
    prompt = (
        "Sos un asistente financiero respondiendo por WhatsApp. Estos son los movimientos del usuario de los "
        f"últimos 6 meses (fecha | tipo | monto | moneda | categoría | descripción):\n\n{lineas}\n\n"
        f'El usuario pregunta: "{pregunta}"\n\n'
        "Respondé en español rioplatense, corto y directo (máximo 3-4 líneas, como un mensaje real de WhatsApp). "
        "Usá solo estos datos, no inventes cifras que no estén ahí. Si no podés responder con esta información, decilo. "
        "No uses markdown ni bullets."
    )
    try:
        client = genai.Client(api_key=os.environ["GEMINI_API_KEY"],
                               http_options=genai_types.HttpOptions(timeout=15000))
        resp = client.models.generate_content(model="gemini-3.6-flash", contents=prompt)
        return (resp.text or "").strip() or None
    except Exception:
        return None

def _wa_a_formato_envio(telefono):
    # Los mensajes ENTRANTES de números argentinos llegan como 549+área+número
    # (formato moderno). El endpoint de ENVÍO de Meta, para Argentina, a veces
    # solo reconoce el formato local viejo 54+área+15+número. Asume área de 2
    # dígitos (CABA/GBA, el único caso de uso real de esta app por ahora).
    if telefono.startswith("549") and len(telefono) == 13:
        return "54" + telefono[3:5] + "15" + telefono[5:]
    return telefono

def _wa_enviar_mensaje(telefono, texto):
    token = os.environ.get("WHATSAPP_TOKEN", "")
    phone_id = os.environ.get("WHATSAPP_PHONE_ID", "")
    if not token or not phone_id:
        print("[whatsapp] falta WHATSAPP_TOKEN o WHATSAPP_PHONE_ID en el entorno")
        return
    try:
        r = req.post(
            f"https://graph.facebook.com/v20.0/{phone_id}/messages",
            headers={"Authorization": f"Bearer {token}"},
            json={
                "messaging_product": "whatsapp",
                "to": _wa_a_formato_envio(telefono),
                "type": "text",
                "text": {"body": texto},
            },
            timeout=10,
        )
        if not r.ok:
            print(f"[whatsapp] error al enviar ({r.status_code}): {r.text}")
    except Exception as e:
        print(f"[whatsapp] excepcion al enviar: {e}")

def _procesar_mensaje_whatsapp(msg):
    wamid = msg.get("id")
    if not wamid or wamid in _wa_procesados:
        return
    _wa_procesados.add(wamid)
    if len(_wa_procesados) > 2000:
        _wa_procesados.clear()

    telefono = msg.get("from", "")
    texto = ((msg.get("text") or {}).get("body") or "").strip()
    print(f"[whatsapp] mensaje recibido de '{telefono}': {texto!r}")
    if not telefono or not texto:
        return

    m = re.match(r"(?i)^vincular\s+(\d{6})$", texto.strip())
    if m:
        codigo = m.group(1)
        entrada = _wa_codigos.get(codigo)
        if not entrada or entrada[1] < time.time():
            _wa_enviar_mensaje(telefono, "❌ Código inválido o vencido. Generá uno nuevo desde la app.")
            return
        db.table("whatsapp_users").upsert({"telefono": telefono, "user_id": entrada[0]}, on_conflict="telefono").execute()
        _wa_codigos.pop(codigo, None)
        _wa_enviar_mensaje(telefono, "✅ ¡Listo! Tu WhatsApp quedó vinculado a tu cuenta de Finanzas.")
        return

    vinculo = db.table("whatsapp_users").select("user_id").eq("telefono", telefono).execute()
    if not vinculo.data:
        _wa_enviar_mensaje(
            telefono,
            "No reconozco este número. Vinculalo primero desde la app: Finanzas → sección WhatsApp → Vincular.",
        )
        return
    user_id = vinculo.data[0]["user_id"]

    parseado = _parse_mensaje_whatsapp(texto)
    if not parseado:
        respuesta = _wa_responder_pregunta(user_id, texto)
        _wa_enviar_mensaje(telefono, respuesta or (
            "No entendí 🤔. Probá algo como: *gasté 500 en supermercado* o *ingreso 300000 sueldo*, "
            "o preguntame algo como *cuánto gasté en comida este mes*."
        ))
        return

    hoy = _hoy_ar().isoformat()
    ok, resultado = _insertar_transaccion(
        user_id, parseado["tipo"], parseado["monto"], hoy,
        categoria=parseado["categoria"], descripcion=parseado["descripcion"], moneda=parseado["moneda"],
    )
    if not ok:
        _wa_enviar_mensaje(
            telefono,
            "⛔ Llegaste al límite de 50 transacciones gratis este mes. Entrá a la app para hacerte Pro y seguir cargando.",
        )
        return

    simbolo = {"ARS": "$", "USD": "USD ", "EUR": "€"}.get(parseado["moneda"], "$")
    emoji = "🔴" if parseado["tipo"] == "Gasto" else "🟢"
    cat_txt = f" · {parseado['categoria']}" if parseado["categoria"] else ""
    desc_txt = f" · {parseado['descripcion']}" if parseado["descripcion"] else ""
    _wa_enviar_mensaje(telefono, f"{emoji} {parseado['tipo']} de {simbolo}{parseado['monto']:,.2f}{cat_txt}{desc_txt} registrado ✓")

@app.route("/api/finanzas/whatsapp/codigo", methods=["POST"])
@login_required
def whatsapp_codigo():
    codigo = f"{random.randint(0, 999999):06d}"
    _wa_codigos[codigo] = (session["user_id"], time.time() + 900)
    numero = re.sub(r"[^0-9]", "", os.environ.get("WHATSAPP_DISPLAY_NUMBER", ""))
    wa_link = f"https://wa.me/{numero}?text={quote(f'VINCULAR {codigo}')}"
    return jsonify({"ok": True, "codigo": codigo, "wa_link": wa_link})

@app.route("/api/finanzas/whatsapp/estado")
@login_required
def whatsapp_estado():
    res = db.table("whatsapp_users").select("telefono").eq("user_id", session["user_id"]).execute()
    if res.data:
        tel = res.data[0]["telefono"]
        return jsonify({"vinculado": True, "telefono_oculto": "•••• " + tel[-4:]})
    return jsonify({"vinculado": False})

@app.route("/api/finanzas/whatsapp/desvincular", methods=["DELETE"])
@login_required
def whatsapp_desvincular():
    db.table("whatsapp_users").delete().eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})

@app.route("/api/finanzas/whatsapp/webhook", methods=["GET"])
def whatsapp_webhook_verificar():
    modo = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge", "")
    if modo == "subscribe" and token and token == os.environ.get("WHATSAPP_VERIFY_TOKEN"):
        return challenge, 200
    return "Forbidden", 403

@app.route("/api/finanzas/whatsapp/webhook", methods=["POST"])
def whatsapp_webhook_recibir():
    app_secret = os.environ.get("WHATSAPP_APP_SECRET", "")
    if app_secret:
        firma = request.headers.get("X-Hub-Signature-256", "")
        esperado = "sha256=" + hmac.new(app_secret.encode(), request.get_data(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(firma, esperado):
            return jsonify({"error": "firma invalida"}), 403

    data = request.get_json(silent=True) or {}
    try:
        for entry in data.get("entry", []):
            for change in entry.get("changes", []):
                for msg in (change.get("value") or {}).get("messages", []):
                    _procesar_mensaje_whatsapp(msg)
    except Exception:
        pass
    return jsonify({"ok": True})

@app.route("/api/finanzas/inflacion")
@login_required
def inflacion():
    global _inflacion_cache
    if time.time() - _inflacion_cache["ts"] < 86400 and _inflacion_cache["data"] is not None:
        return jsonify(_inflacion_cache["data"])
    try:
        r = req.get(
            "https://apis.datos.gob.ar/series/api/series/",
            params={"ids": "148.3_INIVELNAL_DICI_M_26", "limit": 25, "format": "json"},
            timeout=6,
        )
        series = r.json().get("data", [])
        result = [{"mes": row[0][:7], "ipc": row[1]} for row in series if row[1] is not None]
        _inflacion_cache = {"data": result, "ts": time.time()}
        return jsonify(result)
    except Exception:
        return jsonify(_inflacion_cache["data"] or []), 200

def _stats_mes(user_id, mes, anio):
    from collections import defaultdict
    def _norm(s):
        s = (s or "otros").strip().lower()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    inicio_mes = date(anio, mes, 1).isoformat()
    inicio_mes_ant = date(anio if mes > 1 else anio - 1, mes - 1 if mes > 1 else 12, 1).isoformat()
    fin_mes = date(anio if mes < 12 else anio + 1, mes + 1 if mes < 12 else 1, 1).isoformat()
    res_act = db.table("transacciones").select("tipo,monto,moneda,categoria").eq("user_id", user_id).gte("fecha", inicio_mes).lt("fecha", fin_mes).execute()
    res_ant = db.table("transacciones").select("tipo,monto,moneda").eq("user_id", user_id).gte("fecha", inicio_mes_ant).lt("fecha", inicio_mes).execute()
    cotiz = _obtener_cotizaciones()
    def _ars(t):
        conv = _convertir_moneda(float(t["monto"]), t.get("moneda") or "ARS", "ARS", cotiz)
        return conv if conv is not None else 0.0
    cats_total = defaultdict(float)
    cats_label = {}
    gas_act = ing_act = 0.0
    for t in res_act.data:
        m = _ars(t)
        if t["tipo"] == "Gasto":
            key = _norm(t.get("categoria"))
            cats_total[key] += m
            if key not in cats_label:
                cats_label[key] = (t.get("categoria") or "Otros").strip().capitalize()
            gas_act += m
        else:
            ing_act += m
    gas_ant = sum(_ars(t) for t in res_ant.data if t["tipo"] == "Gasto")
    ing_ant  = sum(_ars(t) for t in res_ant.data if t["tipo"] == "Ingreso")
    meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    return {
        "categorias": [{"nombre": cats_label[k], "total": v} for k, v in sorted(cats_total.items(), key=lambda x: -x[1])],
        "resumen": {
            "gastos_actual": gas_act, "ingresos_actual": ing_act,
            "gastos_anterior": gas_ant, "ingresos_anterior": ing_ant,
            "mes_actual": meses[mes - 1], "mes_anterior": meses[(mes - 2) % 12],
        },
        "count_mes": len(res_act.data),
    }

@app.route("/api/finanzas/stats")
@login_required
def stats():
    if not _es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    hoy_ar = _hoy_ar()
    mes  = request.args.get("mes",  type=int, default=hoy_ar.month)
    anio = request.args.get("anio", type=int, default=hoy_ar.year)
    mes  = max(1, min(12, mes))
    return jsonify(_stats_mes(session["user_id"], mes, anio))

_resumen_ia_cache = {}  # (user_id, mes, anio) -> texto ya generado (solo meses cerrados, ver abajo)

@app.route("/api/finanzas/resumen-ia")
@login_required
def resumen_ia():
    if not _es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403
    if not os.environ.get("GEMINI_API_KEY"):
        return jsonify({"ok": False, "error": "ia_no_configurada"}), 503

    hoy_ar = _hoy_ar()
    mes  = request.args.get("mes",  type=int, default=hoy_ar.month)
    anio = request.args.get("anio", type=int, default=hoy_ar.year)
    mes  = max(1, min(12, mes))
    es_mes_actual = (mes == hoy_ar.month and anio == hoy_ar.year)

    cache_key = (session["user_id"], mes, anio)
    if not es_mes_actual and cache_key in _resumen_ia_cache:
        return jsonify({"ok": True, "resumen": _resumen_ia_cache[cache_key]})

    datos = _stats_mes(session["user_id"], mes, anio)
    if datos["count_mes"] == 0:
        return jsonify({"ok": False, "error": "sin_datos"}), 400

    r = datos["resumen"]
    cats_txt = ", ".join(f"{c['nombre']} ${c['total']:.0f}" for c in datos["categorias"][:6]) or "sin gastos"
    prompt = (
        f"Datos financieros de {r['mes_actual']} de un usuario (montos en ARS):\n"
        f"- Ingresos: ${r['ingresos_actual']:.0f} (mes anterior, {r['mes_anterior']}: ${r['ingresos_anterior']:.0f})\n"
        f"- Gastos: ${r['gastos_actual']:.0f} (mes anterior: ${r['gastos_anterior']:.0f})\n"
        f"- Gastos por categoría: {cats_txt}\n\n"
        "Escribí un resumen breve (3-4 oraciones, español rioplatense, tono cercano y directo) sobre cómo viene "
        "el mes comparado con el anterior y en qué se concentra el gasto. Interpretá los números, no los repitas "
        "tal cual. No uses markdown ni bullets, solo texto corrido."
    )
    try:
        client = genai.Client(api_key=os.environ["GEMINI_API_KEY"],
                               http_options=genai_types.HttpOptions(timeout=25000))
        resp = client.models.generate_content(model="gemini-3.6-flash", contents=prompt)
        texto = (resp.text or "").strip()
    except Exception:
        return jsonify({"ok": False, "error": "ia_no_disponible"}), 502

    if not texto:
        return jsonify({"ok": False, "error": "ia_no_disponible"}), 502

    if not es_mes_actual:
        _resumen_ia_cache[cache_key] = texto
    return jsonify({"ok": True, "resumen": texto})

@app.route("/api/finanzas/recurrentes", methods=["GET"])
@login_required
def listar_recurrentes():
    res = db.table("recurrentes").select("*").eq("user_id", session["user_id"]).eq("activo", True).order("creado_en", desc=True).execute()
    return jsonify(res.data)

@app.route("/api/finanzas/recurrentes", methods=["POST"])
@login_required
def crear_recurrente():
    if not _es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403
    try:
        t = request.get_json()
        payload = {
            "tipo":          t["tipo"],
            "monto":         float(t["monto"]),
            "categoria":     t.get("categoria") or "",
            "descripcion":   t.get("descripcion") or "",
            "moneda":        t.get("moneda") or "ARS",
            "frecuencia":    t["frecuencia"],
            "proxima_fecha": t["fecha"],
            "user_id":       session["user_id"],
        }
        res = db.table("recurrentes").insert(payload).execute()
        return jsonify(res.data[0]), 201
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/finanzas/recurrentes/<int:rid>", methods=["DELETE"])
@login_required
def eliminar_recurrente(rid):
    db.table("recurrentes").update({"activo": False}).eq("id", rid).eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})

@app.route("/api/finanzas/<int:tid>", methods=["PUT"])
@login_required
def editar(tid):
    t = request.get_json()
    payload = {
        "tipo":        t["tipo"],
        "monto":       float(t["monto"]),
        "fecha":       t["fecha"],
        "categoria":   t.get("categoria", ""),
        "descripcion": t.get("descripcion", ""),
        "moneda":      t.get("moneda", "ARS"),
        "viaje_id":    t.get("viaje_id"),
    }
    res = db.table("transacciones").update(payload).eq("id", tid).eq("user_id", session["user_id"]).execute()
    return jsonify(res.data[0])

@app.route("/api/finanzas/<int:tid>", methods=["DELETE"])
@login_required
def eliminar(tid):
    db.table("transacciones").delete().eq("id", tid).eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})

@app.route("/api/finanzas/presupuestos", methods=["GET"])
@login_required
def get_presupuestos():
    res = db.table("presupuestos").select("*").eq("user_id", session["user_id"]).execute()
    return jsonify(res.data)

@app.route("/api/finanzas/presupuestos", methods=["POST"])
@login_required
def set_presupuesto():
    d = request.get_json()
    cat    = (d.get("categoria") or "").strip()
    monto  = float(d.get("monto") or 0)
    moneda = d.get("moneda") if d.get("moneda") in ("ARS", "USD", "EUR") else "ARS"
    if not cat or monto <= 0:
        return jsonify({"error": "datos inválidos"}), 400
    db.table("presupuestos").upsert(
        {"user_id": session["user_id"], "categoria": cat, "monto": monto, "moneda": moneda},
        on_conflict="user_id,categoria,moneda"
    ).execute()
    res = (
        db.table("presupuestos").select("*")
        .eq("user_id", session["user_id"]).eq("categoria", cat).eq("moneda", moneda)
        .execute()
    )
    return jsonify(res.data[0] if res.data else {})

@app.route("/api/finanzas/presupuestos/<int:pid>", methods=["DELETE"])
@login_required
def del_presupuesto(pid):
    db.table("presupuestos").delete().eq("id", pid).eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})

# ── Viajes ──────────────────────────────────────────────────────────────────

def _payload_viaje(d):
    return {
        "nombre":       (d.get("nombre") or "").strip(),
        "fecha_inicio": d.get("fecha_inicio"),
        "fecha_fin":    d.get("fecha_fin"),
    }

@app.route("/api/finanzas/viajes", methods=["GET"])
@login_required
def listar_viajes():
    if not _es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    viajes = (
        db.table("viajes").select("*").eq("user_id", session["user_id"])
        .order("fecha_inicio", desc=True).execute().data
    )
    ids = [v["id"] for v in viajes]
    gastos = {}
    if ids:
        cotiz = _obtener_cotizaciones()
        txs = (
            db.table("transacciones").select("viaje_id,tipo,monto,moneda")
            .eq("user_id", session["user_id"]).in_("viaje_id", ids).execute().data
        )
        for t in txs:
            if t["tipo"] != "Gasto":
                continue
            vid = t["viaje_id"]
            origen = t.get("moneda") or "ARS"
            monto = float(t["monto"])
            gastos.setdefault(vid, {"ARS": 0.0, "USD": 0.0, "EUR": 0.0})
            for destino in ("ARS", "USD", "EUR"):
                conv = _convertir_moneda(monto, origen, destino, cotiz)
                if conv is not None:
                    gastos[vid][destino] += conv
    for v in viajes:
        v["gastado"] = gastos.get(v["id"], {"ARS": 0.0, "USD": 0.0, "EUR": 0.0})
    return jsonify(viajes)

@app.route("/api/finanzas/viajes", methods=["POST"])
@login_required
def crear_viaje():
    if not _es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403
    d = request.get_json()
    payload = _payload_viaje(d)
    if not payload["nombre"] or not payload["fecha_inicio"] or not payload["fecha_fin"]:
        return jsonify({"ok": False, "error": "missing_trip_fields"}), 400
    payload["user_id"] = session["user_id"]
    res = db.table("viajes").insert(payload).execute()
    return jsonify(res.data[0]), 201

@app.route("/api/finanzas/viajes/<int:vid>", methods=["PUT"])
@login_required
def editar_viaje(vid):
    if not _es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403
    d = request.get_json()
    payload = _payload_viaje(d)
    if not payload["nombre"] or not payload["fecha_inicio"] or not payload["fecha_fin"]:
        return jsonify({"ok": False, "error": "missing_trip_fields"}), 400
    res = db.table("viajes").update(payload).eq("id", vid).eq("user_id", session["user_id"]).execute()
    if not res.data:
        return jsonify({"ok": False, "error": "no encontrado"}), 404
    return jsonify(res.data[0])

@app.route("/api/finanzas/viajes/<int:vid>", methods=["DELETE"])
@login_required
def eliminar_viaje(vid):
    db.table("viajes").delete().eq("id", vid).eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})

@app.route("/api/finanzas/viajes/<int:vid>", methods=["GET"])
@login_required
def detalle_viaje(vid):
    if not _es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    viaje = db.table("viajes").select("*").eq("id", vid).eq("user_id", session["user_id"]).execute().data
    if not viaje:
        return jsonify({"error": "no encontrado"}), 404
    viaje = viaje[0]

    txs = (
        db.table("transacciones").select("*")
        .eq("user_id", session["user_id"]).eq("viaje_id", vid)
        .order("fecha").execute().data
    )

    cotiz = _obtener_cotizaciones()
    totales = {"ARS": 0.0, "USD": 0.0, "EUR": 0.0}
    por_cat = {"ARS": {}, "USD": {}, "EUR": {}}
    for t in txs:
        origen = t.get("moneda") or "ARS"
        monto = float(t["monto"])
        t["monto_usd"] = _convertir_moneda(monto, origen, "USD", cotiz)
        if t["tipo"] != "Gasto":
            continue
        cat = (t.get("categoria") or "").strip() or "Sin categoría"
        for destino in ("ARS", "USD", "EUR"):
            conv = _convertir_moneda(monto, origen, destino, cotiz)
            if conv is None:
                continue
            totales[destino] += conv
            por_cat[destino][cat] = por_cat[destino].get(cat, 0.0) + conv

    por_cat_lista = {
        m: sorted([{"nombre": c, "total": v} for c, v in cats.items()], key=lambda x: -x["total"])
        for m, cats in por_cat.items()
    }

    return jsonify({
        "viaje": viaje, "transacciones": txs, "totales": totales,
        "por_categoria": por_cat_lista,
    })

@app.route("/api/finanzas/suscribir")
@login_required
def suscribir():
    mp_token = os.environ.get("MP_ACCESS_TOKEN", "")
    base_url = os.environ.get("BASE_URL") or request.host_url.rstrip("/")
    r = req.post(
        "https://api.mercadopago.com/preapproval",
        headers={"Authorization": f"Bearer {mp_token}"},
        json={
            "reason": "Finanzas Pro",
            "external_reference": session["user_id"],
            "payer_email": f"{session['username']}@finanzas.local",
            "auto_recurring": {
                "frequency": 1,
                "frequency_type": "months",
                "transaction_amount": 1999,
                "currency_id": "ARS",
            },
            "back_url": f"{base_url}/finanzas",
            "notification_url": f"{base_url}/api/finanzas/webhook/mercadopago",
        }
    )
    data = r.json()
    init_point = data.get("init_point")
    if not init_point:
        return jsonify({"error": "No se pudo crear la suscripción", "detalle": data}), 500
    return redirect(init_point)

@app.route("/api/finanzas/webhook/mercadopago", methods=["POST"])
def webhook_mp():
    data = request.get_json(silent=True) or {}
    topic = data.get("type") or request.args.get("topic", "")
    resource_id = data.get("data", {}).get("id") or request.args.get("id", "")

    if topic in ("subscription_preapproval", "preapproval") and resource_id:
        mp_token = os.environ.get("MP_ACCESS_TOKEN", "")
        r = req.get(
            f"https://api.mercadopago.com/preapproval/{resource_id}",
            headers={"Authorization": f"Bearer {mp_token}"}
        )
        preapproval = r.json()
        if preapproval.get("status") == "authorized":
            user_id = preapproval.get("external_reference")
            if user_id:
                db.table("usuarios").update({"plan": "pro"}).eq("id", user_id).execute()

    return jsonify({"ok": True})

if __name__ == "__main__":
    import socket
    port = int(os.environ.get("PORT", 5001))
    ip = socket.gethostbyname(socket.gethostname())
    print(f"\n  Local:  http://localhost:{port}/finanzas\n")
    app.run(host="0.0.0.0", port=port, debug=False)
