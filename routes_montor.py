"""Transacciones, export, import, presupuestos, recurrentes, stats, resumen IA."""
import csv
import io
import os
from datetime import date, timedelta
from flask import Blueprint, Response, jsonify, request, session

import finanzas_server as core

bp = Blueprint("montor", __name__)


@bp.route("/api/finanzas/cotizaciones")
def cotizaciones():
    return jsonify(core._obtener_cotizaciones())


@bp.route("/api/finanzas", methods=["GET"])
@core.login_required
def listar():
    try:
        core._procesar_recurrentes(session["user_id"])
    except Exception:
        pass
    es_pro = core._es_pro(session["user_id"])
    query = core.db.table("transacciones").select("*").eq("user_id", session["user_id"])
    if not es_pro:
        desde = (core._hoy_ar() - timedelta(days=90)).isoformat()
        query = query.gte("fecha", desde)
    res = query.order("fecha", desc=True).execute()
    return jsonify(res.data)


@bp.route("/api/finanzas", methods=["POST"])
@core.login_required
def agregar():
    t = request.get_json()
    ok, resultado = core._insertar_transaccion(
        session["user_id"], t["tipo"], t["monto"], t["fecha"],
        t.get("categoria", ""), t.get("descripcion", ""), t.get("moneda", "ARS"),
        t.get("viaje_id"),
    )
    if not ok:
        return jsonify({"ok": False, "error": resultado}), 403
    return jsonify(resultado), 201


@bp.route("/api/finanzas/export", methods=["GET"])
@core.login_required
def exportar():
    if not core._es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    res = core.db.table("transacciones").select("*").eq("user_id", session["user_id"]).order("fecha", desc=True).execute()
    output = io.StringIO()
    output.write("sep=;\n")
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Fecha", "Tipo", "Monto", "Moneda", "Categoria", "Descripcion"])
    for t in res.data:
        writer.writerow([
            t["fecha"], t["tipo"], t["monto"],
            t.get("moneda","ARS"), t.get("categoria",""), t.get("descripcion",""),
        ])
    encoded = output.getvalue().encode("utf-8-sig")
    return Response(encoded, mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=montor.csv"})


@bp.route("/api/finanzas/export/excel", methods=["GET"])
@core.login_required
def exportar_excel():
    if not core._es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    res = core.db.table("transacciones").select("*").eq("user_id", session["user_id"]).order("fecha", desc=True).execute()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    headers = ["Fecha", "Tipo", "Monto", "Moneda", "Categoría", "Descripción"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for t in res.data:
        ws.append([
            t["fecha"], t["tipo"], float(t["monto"]),
            t.get("moneda", "ARS"), t.get("categoria", ""), t.get("descripcion", ""),
        ])
    for col, width in zip("ABCDEF", (12, 10, 14, 8, 20, 34)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=montor.xlsx"},
    )


@bp.route("/api/finanzas/export/pdf", methods=["GET"])
@core.login_required
def exportar_pdf():
    if not core._es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    res = core.db.table("transacciones").select("*").eq("user_id", session["user_id"]).order("fecha", desc=True).execute()

    from fpdf import FPDF

    def _safe(s):
        return (s or "").encode("latin-1", "replace").decode("latin-1")

    total_gastos = sum(float(t["monto"]) for t in res.data if t["tipo"] == "Gasto")
    total_ingresos = sum(float(t["monto"]) for t in res.data if t["tipo"] == "Ingreso")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _safe("Montor - Historial de transacciones"), ln=1)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, _safe(f"Generado el {core._hoy_ar().isoformat()} - Usuario: {session['username']}"), ln=1)
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, _safe(f"Total ingresos: {total_ingresos:,.2f}   Total gastos: {total_gastos:,.2f}"), ln=1)
    pdf.ln(4)

    col_widths = [22, 18, 26, 14, 40, 68]
    headers = ["Fecha", "Tipo", "Monto", "Moneda", "Categoria", "Descripcion"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(22, 163, 74)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 7, h, border=1, fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for t in res.data:
        fila = [
            t["fecha"], t["tipo"], f'{float(t["monto"]):,.2f}',
            t.get("moneda", "ARS"), (t.get("categoria") or "")[:22], (t.get("descripcion") or "")[:45],
        ]
        for w, val in zip(col_widths, fila):
            pdf.cell(w, 6, _safe(str(val)), border=1)
        pdf.ln()
        if pdf.get_y() > 270:
            pdf.add_page()

    salida = bytes(pdf.output())
    return Response(
        salida, mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=montor.pdf"},
    )


@bp.route("/api/finanzas/import/preview", methods=["POST"])
@core.login_required
def import_preview():
    if not core._es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403

    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename.lower().endswith(".csv"):
        return jsonify({"ok": False, "error": "invalid_file_type"}), 400

    raw = archivo.read(2_000_000)
    texto = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        return jsonify({"ok": False, "error": "cant_read_file"}), 400

    muestra = texto[:2000]
    try:
        delim = csv.Sniffer().sniff(muestra, delimiters=",;\t").delimiter
    except csv.Error:
        delim = ";" if muestra.count(";") > muestra.count(",") else ","

    filas = list(csv.reader(io.StringIO(texto), delimiter=delim))
    if len(filas) < 2:
        return jsonify({"ok": False, "error": "empty_file"}), 400

    headers = [core._norm_header(h) for h in filas[0]]
    idx_fecha = core._buscar_col(headers, ["fecha", "date"])
    idx_monto = core._buscar_col(headers, ["importe", "monto", "valor", "amount", "total"])
    idx_desc  = core._buscar_col(headers, ["descripcion", "concepto", "detalle", "description", "glosa"])
    idx_debe  = core._buscar_col(headers, ["debe", "egreso", "cargo", "debito"])
    idx_haber = core._buscar_col(headers, ["haber", "ingreso", "abono", "credito"])

    if idx_fecha is None or (idx_monto is None and (idx_debe is None or idx_haber is None)):
        return jsonify({
            "ok": False,
            "error": "unrecognized_columns",
        }), 400

    transacciones = []
    errores = 0
    for fila in filas[1:]:
        if not fila or all(not c.strip() for c in fila):
            continue
        try:
            fecha = core._parse_fecha_import(fila[idx_fecha]) if idx_fecha < len(fila) else None
            desc  = fila[idx_desc].strip() if idx_desc is not None and idx_desc < len(fila) else ""

            if idx_monto is not None:
                monto_raw = core._parse_monto_import(fila[idx_monto]) if idx_monto < len(fila) else None
                if not monto_raw:
                    errores += 1
                    continue
                tipo, monto = ("Gasto" if monto_raw < 0 else "Ingreso"), abs(monto_raw)
            else:
                debe  = abs(core._parse_monto_import(fila[idx_debe])  or 0) if idx_debe  < len(fila) else 0
                haber = abs(core._parse_monto_import(fila[idx_haber]) or 0) if idx_haber < len(fila) else 0
                if haber > 0:
                    tipo, monto = "Ingreso", haber
                elif debe > 0:
                    tipo, monto = "Gasto", debe
                else:
                    errores += 1
                    continue

            if not fecha:
                errores += 1
                continue

            transacciones.append({
                "fecha": fecha, "tipo": tipo, "monto": round(monto, 2),
                "descripcion": desc, "categoria": "", "moneda": "ARS",
            })
        except Exception:
            errores += 1

        if len(transacciones) >= 1000:
            break

    return jsonify({"ok": True, "transacciones": transacciones, "errores": errores})


@bp.route("/api/finanzas/import/confirm", methods=["POST"])
@core.login_required
def import_confirm():
    if not core._es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403

    data = request.get_json() or {}
    filas = data.get("transacciones") or []
    if not isinstance(filas, list) or not filas:
        return jsonify({"ok": False, "error": "nothing_to_import"}), 400
    if len(filas) > 1000:
        return jsonify({"ok": False, "error": "too_many_transactions"}), 400

    payload = []
    for t in filas:
        try:
            fecha = date.fromisoformat(t["fecha"]).isoformat()
            # Un CSV importado suele traer movimientos de fechas pasadas: se
            # usa la cotización histórica de cada fecha, no la de hoy.
            cotiz = core._cotizacion_para_fecha(fecha)
            payload.append({
                "tipo":        "Ingreso" if t.get("tipo") == "Ingreso" else "Gasto",
                "monto":       abs(float(t["monto"])),
                "fecha":       fecha,
                "categoria":   (t.get("categoria") or "").strip()[:60],
                "descripcion": (t.get("descripcion") or "").strip()[:200],
                "moneda":      t.get("moneda") if t.get("moneda") in ("ARS", "USD", "EUR") else "ARS",
                "user_id":     session["user_id"],
                "cotizacion_usd": cotiz.get("USD"),
                "cotizacion_eur": cotiz.get("EUR"),
            })
        except (KeyError, ValueError, TypeError):
            continue

    if not payload:
        return jsonify({"ok": False, "error": "nothing_valid_to_import"}), 400

    insertados = 0
    for i in range(0, len(payload), 500):
        lote = payload[i:i + 500]
        core.db.table("transacciones").insert(lote).execute()
        insertados += len(lote)

    return jsonify({"ok": True, "insertados": insertados})


@bp.route("/api/finanzas/stats")
@core.login_required
def stats():
    if not core._es_pro(session["user_id"]):
        return jsonify({"error": "pro_requerido"}), 403
    hoy_ar = core._hoy_ar()
    mes  = request.args.get("mes",  type=int, default=hoy_ar.month)
    anio = request.args.get("anio", type=int, default=hoy_ar.year)
    mes  = max(1, min(12, mes))
    return jsonify(core._stats_mes(session["user_id"], mes, anio))


@bp.route("/api/finanzas/resumen-ia")
@core.login_required
def resumen_ia():
    if not core._es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403
    if not os.environ.get("GEMINI_API_KEY"):
        return jsonify({"ok": False, "error": "ia_no_configurada"}), 503

    hoy_ar = core._hoy_ar()
    mes  = request.args.get("mes",  type=int, default=hoy_ar.month)
    anio = request.args.get("anio", type=int, default=hoy_ar.year)
    mes  = max(1, min(12, mes))
    lang = "en" if request.args.get("lang") == "en" else "es"
    es_mes_actual = (mes == hoy_ar.month and anio == hoy_ar.year)

    cache_key = (session["user_id"], mes, anio, lang)
    if not es_mes_actual and cache_key in core._resumen_ia_cache:
        return jsonify({"ok": True, "resumen": core._resumen_ia_cache[cache_key]})

    datos = core._stats_mes(session["user_id"], mes, anio)
    if datos["count_mes"] == 0:
        return jsonify({"ok": False, "error": "sin_datos"}), 400

    r = datos["resumen"]
    cats_txt = ", ".join(f"{c['nombre']} ${c['total']:.0f}" for c in datos["categorias"][:6]) or "sin gastos"
    if lang == "en":
        instruccion_idioma = (
            "Write your answer in English. The month names in the data below are in Spanish — translate them "
            "into English in your answer (e.g. 'Agosto' -> 'August')."
        )
    else:
        instruccion_idioma = "Escribí tu respuesta en español rioplatense."
    prompt = (
        f"Datos financieros de {r['mes_actual']} de un usuario (montos en ARS):\n"
        f"- Ingresos: ${r['ingresos_actual']:.0f} (mes anterior, {r['mes_anterior']}: ${r['ingresos_anterior']:.0f})\n"
        f"- Gastos: ${r['gastos_actual']:.0f} (mes anterior: ${r['gastos_anterior']:.0f})\n"
        f"- Gastos por categoría: {cats_txt}\n\n"
        f"Escribí un resumen breve (3-4 oraciones, tono cercano y directo) sobre cómo viene el mes comparado con "
        "el anterior y en qué se concentra el gasto. Interpretá los números, no los repitas tal cual. No uses "
        f"markdown ni bullets, solo texto corrido. {instruccion_idioma}"
    )
    try:
        client = core.genai.Client(api_key=os.environ["GEMINI_API_KEY"],
                                    http_options=core.genai_types.HttpOptions(timeout=25000))
        resp = client.models.generate_content(model="gemini-3.5-flash-lite", contents=prompt)
        texto = (resp.text or "").strip()
    except Exception:
        return jsonify({"ok": False, "error": "ia_no_disponible"}), 502

    if not texto:
        return jsonify({"ok": False, "error": "ia_no_disponible"}), 502

    if not es_mes_actual:
        core._resumen_ia_cache[cache_key] = texto
    return jsonify({"ok": True, "resumen": texto})


@bp.route("/api/finanzas/recurrentes", methods=["GET"])
@core.login_required
def listar_recurrentes():
    res = core.db.table("recurrentes").select("*").eq("user_id", session["user_id"]).eq("activo", True).order("creado_en", desc=True).execute()
    return jsonify(res.data)


@bp.route("/api/finanzas/recurrentes", methods=["POST"])
@core.login_required
def crear_recurrente():
    if not core._es_pro(session["user_id"]):
        return jsonify({"ok": False, "error": "pro_requerido"}), 403
    try:
        t = request.get_json()
        payload = {
            "tipo":          t["tipo"],
            "monto":         float(t["monto"]),
            "categoria":     t.get("categoria") or "",
            "descripcion":   t.get("descripcion") or "",
            "moneda":        t.get("moneda") or "ARS",
            "frecuencia":    t["frecuencia"],
            "proxima_fecha": t["fecha"],
            "user_id":       session["user_id"],
        }
        res = core.db.table("recurrentes").insert(payload).execute()
        return jsonify(res.data[0]), 201
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/finanzas/recurrentes/<int:rid>", methods=["DELETE"])
@core.login_required
def eliminar_recurrente(rid):
    core.db.table("recurrentes").update({"activo": False}).eq("id", rid).eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})


@bp.route("/api/finanzas/<int:tid>", methods=["PUT"])
@core.login_required
def editar(tid):
    t = request.get_json()
    payload = {
        "tipo":        t["tipo"],
        "monto":       float(t["monto"]),
        "fecha":       t["fecha"],
        "categoria":   t.get("categoria", ""),
        "descripcion": t.get("descripcion", ""),
        "moneda":      t.get("moneda", "ARS"),
        "viaje_id":    t.get("viaje_id"),
    }
    # Si la fila es de antes de que existiera la foto de cotización (o nunca
    # se pudo tomar), se completa ahora con la cotización histórica de su
    # fecha - pero nunca se pisa una ya guardada, para no revaluar una
    # transacción vieja con el tipo de cambio de hoy.
    actual = core.db.table("transacciones").select("cotizacion_usd,cotizacion_eur").eq("id", tid).eq("user_id", session["user_id"]).execute().data
    if actual and actual[0].get("cotizacion_usd") is None:
        cotiz = core._cotizacion_para_fecha(payload["fecha"])
        payload["cotizacion_usd"] = cotiz.get("USD")
        payload["cotizacion_eur"] = cotiz.get("EUR")
    res = core.db.table("transacciones").update(payload).eq("id", tid).eq("user_id", session["user_id"]).execute()
    return jsonify(res.data[0])


@bp.route("/api/finanzas/<int:tid>", methods=["DELETE"])
@core.login_required
def eliminar(tid):
    core.db.table("transacciones").delete().eq("id", tid).eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})


@bp.route("/api/finanzas/presupuestos", methods=["GET"])
@core.login_required
def get_presupuestos():
    res = core.db.table("presupuestos").select("*").eq("user_id", session["user_id"]).execute()
    return jsonify(res.data)


@bp.route("/api/finanzas/presupuestos", methods=["POST"])
@core.login_required
def set_presupuesto():
    d = request.get_json()
    cat    = (d.get("categoria") or "").strip()
    monto  = float(d.get("monto") or 0)
    moneda = d.get("moneda") if d.get("moneda") in ("ARS", "USD", "EUR") else "ARS"
    if not cat or monto <= 0:
        return jsonify({"error": "datos inválidos"}), 400
    core.db.table("presupuestos").upsert(
        {"user_id": session["user_id"], "categoria": cat, "monto": monto, "moneda": moneda},
        on_conflict="user_id,categoria,moneda"
    ).execute()
    res = (
        core.db.table("presupuestos").select("*")
        .eq("user_id", session["user_id"]).eq("categoria", cat).eq("moneda", moneda)
        .execute()
    )
    return jsonify(res.data[0] if res.data else {})


@bp.route("/api/finanzas/presupuestos/<int:pid>", methods=["DELETE"])
@core.login_required
def del_presupuesto(pid):
    core.db.table("presupuestos").delete().eq("id", pid).eq("user_id", session["user_id"]).execute()
    return jsonify({"ok": True})


@bp.route("/api/finanzas/inflacion")
@core.login_required
def inflacion():
    if core.time.time() - core._inflacion_cache["ts"] < 86400 and core._inflacion_cache["data"] is not None:
        return jsonify(core._inflacion_cache["data"])
    try:
        r = core.req.get(
            "https://apis.datos.gob.ar/series/api/series/",
            params={"ids": "148.3_INIVELNAL_DICI_M_26", "limit": 25, "format": "json"},
            timeout=6,
        )
        series = r.json().get("data", [])
        result = [{"mes": row[0][:7], "ipc": row[1]} for row in series if row[1] is not None]
        core._inflacion_cache = {"data": result, "ts": core.time.time()}
        return jsonify(result)
    except Exception:
        return jsonify(core._inflacion_cache["data"] or []), 200
